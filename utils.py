import re
import time
import random
from datetime import datetime, timedelta
import os
from pathlib import Path
import signal
from functools import wraps

from openpyxl import load_workbook
import pandas as pd
from logger import setup_logger


# 请求头集合
MyHeaders = [
    "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",
    "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
    'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
    'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
    'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7",
    "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "
]



def timeout(seconds):
    def decorator(func):
        def handler(signum, frame):
            raise TimeoutError("Function timed out")

        @wraps(func)
        def wrapper(*args, **kwargs):
            signal.signal(signal.SIGALRM, handler)
            signal.alarm(seconds)
            try:
                result = func(*args, **kwargs)
            finally:
                signal.alarm(0)
            return result
        return wrapper
    return decorator


def read_key_words(file_path="key_words.md"):
    key_list = []
    f = open(file_path, encoding='utf-8')  # 返回一个文件对象
    lines = f.readlines()  # 读取全部内容
    for line in lines:
        # 如果读到空行，就跳过
        if line.isspace():
            continue
        else:
            # 去除文本中的换行等等，可以追加其他操作
            line = line.replace("\n", "")
            line = line.replace("\t", "")
            # 处理完成后的行，追加到列表中
            key_list.append(line)
    f.close()
    return key_list


def get_date(yesterday=False): 
    # 获取当前日期
    date = datetime.now()

    # 计算昨天的日期
    if yesterday:
        date = date - timedelta(days=1)
    
    # 将日期格式化为字符串
    return date.strftime("%Y-%m-%d"), date.strftime("%Y年%m月")


def append_sheet_to_excel(data_frame, file_path, sheet_name):
    try:
        # 尝试加载已存在的Excel文件
        book = load_workbook(file_path)
        
        if sheet_name in book.sheetnames:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                startrow = writer.sheets[sheet_name].max_row
                data_frame.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        # 如果Excel文件不存在，则创建新的文件和sheet页
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)


# 创建当前日期文件夹
now, year_month = get_date(yesterday=False)
os.makedirs(f'xls_files/{year_month}', exist_ok=True)

# 生成随机睡眠时间
sleep_time = random.uniform(1, 2)

# 设置日志对象
os.makedirs(f'logs', exist_ok=True)
logger = setup_logger(
    log_name='Spider System',
    log_filename=Path('logs/spider.log'),
    log_level='info',
    use_console=False
)